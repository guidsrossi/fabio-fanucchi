import glob
import json
import os
import re
import unicodedata
from pathlib import Path

import openpyxl

DOWNLOADS = Path(r'C:\Users\guids\Downloads')
DESTINO_JSON = Path(__file__).resolve().parents[1] / 'data' / 'notas-primeiro-bimestre.json'
BANCO = DOWNLOADS / 'desenvolvimento.xlsx'
ABA = 'notas_bimestrais'
HEADERS = ['id', 'ano', 'bimestre', 'nome', 'turma', 'situacao', 'frequencia', 'disciplina', 'nota', 'fonte']


def normalizar(valor):
    texto = unicodedata.normalize('NFD', str(valor or '')).encode('ascii', 'ignore').decode().upper()
    return ' '.join(re.sub(r'[^A-Z0-9 ]+', ' ', texto).split())


def turma_da_descricao(valor):
    encontrados = re.findall(r'([123])\s+SERIE\s+([A-E])', normalizar(valor))
    return ''.join(encontrados[-1]) if encontrados else ''


def numero(valor):
    if valor in (None, '', '-'): return None
    try:
        resultado = float(str(valor).replace(',', '.'))
        return int(resultado) if resultado.is_integer() else resultado
    except (TypeError, ValueError):
        return None


def extrair_planilha(caminho):
    ws = openpyxl.load_workbook(caminho, data_only=True, read_only=True).active
    turma = turma_da_descricao(ws['B6'].value)
    total_aulas = numero(ws['B8'].value) or 0
    cabecalhos = [celula.value for celula in ws[11]]
    subcabecalhos = [celula.value for celula in ws[12]]
    disciplinas = []
    for coluna, subcabecalho in enumerate(subcabecalhos, 1):
        if normalizar(subcabecalho) != 'M': continue
        inicio = coluna
        while inicio >= 1 and not cabecalhos[inicio - 1]: inicio -= 1
        disciplina = str(cabecalhos[inicio - 1] or '').split('\n')[0].strip()
        disciplinas.append((coluna, disciplina))

    estudantes = []
    for linha in ws.iter_rows(min_row=13, values_only=True):
        nome = str(linha[0] or '').strip()
        situacao = str(linha[1] or '').strip()
        if not nome or normalizar(situacao) != 'ATIVO': continue
        notas = []
        faltas = 0
        for coluna, disciplina in disciplinas:
            nota = numero(linha[coluna - 1] if coluna - 1 < len(linha) else None)
            falta = numero(linha[coluna] if coluna < len(linha) else None)
            if falta is not None: faltas += falta
            if nota is not None and 0 <= nota <= 10:
                notas.append({'disciplina': disciplina, 'nota': nota, 'bimestre': 1})
        frequencia = round(max(0, (total_aulas - faltas) / total_aulas), 4) if total_aulas else ''
        estudantes.append({
            'id': f'{turma}-{len(estudantes) + 1}', 'nome': nome, 'turma': turma,
            'serie': turma, 'situacao': situacao, 'frequencia': frequencia, 'notas': notas,
            'fonte': os.path.basename(caminho),
        })
    return estudantes


arquivos = sorted(glob.glob(str(DOWNLOADS / 'MAPAO*PRIMEIRO_BIMESTRE*.xlsx')))
estudantes = [estudante for arquivo in arquivos for estudante in extrair_planilha(arquivo)]
DESTINO_JSON.parent.mkdir(parents=True, exist_ok=True)
DESTINO_JSON.write_text(json.dumps(estudantes, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')

banco = openpyxl.load_workbook(BANCO)
if ABA in banco.sheetnames:
    del banco[ABA]
ws = banco.create_sheet(ABA)
ws.append(HEADERS)
registro_id = 1
for estudante in estudantes:
    for nota in estudante['notas']:
        ws.append([
            registro_id, 2026, 1, estudante['nome'], estudante['turma'], estudante['situacao'],
            estudante['frequencia'], nota['disciplina'], nota['nota'], estudante['fonte'],
        ])
        registro_id += 1
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
banco.save(BANCO)

print(json.dumps({
    'arquivos': len(arquivos), 'estudantes': len(estudantes),
    'notas': sum(len(item['notas']) for item in estudantes),
    'turmas': sorted({item['turma'] for item in estudantes}),
    'json': str(DESTINO_JSON), 'banco': str(BANCO),
}, ensure_ascii=False))
