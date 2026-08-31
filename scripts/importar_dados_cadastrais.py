import json
import re
import unicodedata
from collections import Counter
from datetime import datetime
from difflib import SequenceMatcher
from html.parser import HTMLParser
from pathlib import Path

import openpyxl

DOWNLOADS = Path(r'C:\Users\guids\Downloads')
BANCO = DOWNLOADS / 'desenvolvimento.xlsx'
DESTINO = Path(__file__).resolve().parents[1] / 'data' / 'dados-cadastrais-estudantes.json'


class LeitorTabela(HTMLParser):
    def __init__(self):
        super().__init__(); self.rows = []; self.row = None; self.cell = None
    def handle_starttag(self, tag, attrs):
        if tag == 'tr': self.row = []
        if tag in ('td', 'th'): self.cell = ''
    def handle_data(self, data):
        if self.cell is not None: self.cell += data
    def handle_endtag(self, tag):
        if tag in ('td', 'th') and self.cell is not None:
            self.row.append(' '.join(self.cell.split())); self.cell = None
        if tag == 'tr' and self.row is not None:
            self.rows.append(self.row); self.row = None


def normalizar(valor):
    texto = unicodedata.normalize('NFD', str(valor or '')).encode('ascii', 'ignore').decode().upper()
    return ' '.join(re.sub(r'[^A-Z0-9 ]+', ' ', texto).split())


def similaridade(a, b):
    return SequenceMatcher(None, normalizar(a), normalizar(b)).ratio()


def data_iso(valor):
    try: return datetime.strptime(str(valor), '%d/%m/%Y').strftime('%Y-%m-%d')
    except ValueError: return ''


banco = openpyxl.load_workbook(BANCO)
usuarios = banco['usuarios']
cabecalhos = [str(c.value or '').strip() for c in usuarios[1]]
for cabecalho in ['ra', 'data_nascimento']:
    if cabecalho not in cabecalhos:
        usuarios.cell(1, len(cabecalhos) + 1, cabecalho); cabecalhos.append(cabecalho)
indice_ra = cabecalhos.index('ra') + 1
indice_nascimento = cabecalhos.index('data_nascimento') + 1

alunos_banco = []
turmas_por_nome = {}
for numero_linha, linha in enumerate(usuarios.iter_rows(min_row=2, values_only=True), 2):
    if str(linha[4] or '').strip().lower() != 'estudante': continue
    aluno = {'linha': numero_linha, 'id': str(linha[0]), 'nome': str(linha[1] or ''), 'turma': str(linha[5] or '').strip().upper()}
    alunos_banco.append(aluno); turmas_por_nome.setdefault(normalizar(aluno['nome']), []).append(aluno['turma'])

arquivos = []
for caminho in sorted(DOWNLOADS.glob('Matrícula - Relação de Alunos por Classe*.htm.html')):
    leitor = LeitorTabela(); leitor.feed(caminho.read_text(encoding='utf-8'))
    registros = []
    for linha in leitor.rows[1:]:
        if len(linha) < 8: continue
        registros.append({'nome': linha[3], 'ra': f'{linha[4]}-{linha[5]}/{linha[6]}', 'data_nascimento': data_iso(linha[7])})
    contagem = Counter(turma for registro in registros for turma in turmas_por_nome.get(normalizar(registro['nome']), []))
    turma = contagem.most_common(1)[0][0] if contagem else ''
    arquivos.append({'arquivo': caminho.name, 'turma': turma, 'registros': registros})

por_turma = {item['turma']: item['registros'] for item in arquivos if item['turma']}
resultado = []
nao_encontrados = []
for aluno in alunos_banco:
    candidatos = por_turma.get(aluno['turma'], [])
    exatos = [item for item in candidatos if normalizar(item['nome']) == normalizar(aluno['nome'])]
    encontrado = exatos[0] if len(exatos) == 1 else None
    if not encontrado:
        palavras_aluno = set(normalizar(aluno['nome']).split())
        por_conjunto = [item for item in candidatos if len(palavras_aluno) >= 2 and palavras_aluno.issubset(set(normalizar(item['nome']).split()))]
        if len(por_conjunto) == 1: encontrado = por_conjunto[0]
    if not encontrado:
        pontuados = sorted(((similaridade(aluno['nome'], item['nome']), item) for item in candidatos), key=lambda item: item[0], reverse=True)
        if pontuados and pontuados[0][0] >= 0.78 and (len(pontuados) == 1 or pontuados[0][0] - pontuados[1][0] >= 0.08):
            encontrado = pontuados[0][1]
    if not encontrado:
        nao_encontrados.append({'nome': aluno['nome'], 'turma': aluno['turma']}); continue
    usuarios.cell(aluno['linha'], indice_ra, encontrado['ra'])
    usuarios.cell(aluno['linha'], indice_nascimento, encontrado['data_nascimento'])
    resultado.append({**aluno, 'ra': encontrado['ra'], 'data_nascimento': encontrado['data_nascimento']})

banco.save(BANCO)
DESTINO.parent.mkdir(parents=True, exist_ok=True)
DESTINO.write_text(json.dumps(resultado, ensure_ascii=False, separators=(',', ':')), encoding='utf-8')
print(json.dumps({'arquivos': len(arquivos), 'atualizados': len(resultado), 'nao_encontrados': len(nao_encontrados), 'turmas': {item['arquivo']: item['turma'] for item in arquivos}, 'amostra_nao_encontrados': nao_encontrados[:20]}, ensure_ascii=False))
